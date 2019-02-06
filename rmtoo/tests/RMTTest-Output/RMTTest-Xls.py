# (c) 2018 Kristoffer Nordstroem, see COPYING

import os
from datetime import datetime
import openpyxl
import pytest
try:
    from unittest.mock import Mock  # py33 ff.
except ImportError:
    from mock import Mock  # py27

from rmtoo.outputs.xls import XlsHandler as xh
from rmtoo.lib.Requirement import Requirement
from rmtoo.tests.lib.Utils import create_tmp_dir, delete_tmp_dir
from rmtoo.tests.lib.TestConfig import TestConfig

LDIR = os.path.dirname(os.path.abspath(__file__))


def create_req(req_id):
    req_txt_items = {}
    inv_date = datetime.strptime('1970-01-01', "%Y-%m-%d").date()
    req_values = {
        'Name': 'Aston Martin DB5',
        'Topic': 'Escape Routes',
        'Description': "Flying out the roof",
        'Priority': 'development',
        'Owner': '007',
        'Invented on': inv_date.isoformat(),
        'Invented by': 'Q',
        'Status': 'not done',
        'Class': 'requirement'
    }
    req = Requirement(u"", req_id, None, None, TestConfig())
    for key, value in req_values.items():
        req.record.append(MockRecordEntry(key, value))
        req_txt_items[key] = MockRecordEntry(None, value)
    req_txt_items['Invented on'] = inv_date
    req.values.update(req_txt_items)
    return req


class MockRecordEntry:
    def __init__(self, tag, txt_str):
        self.__txt_tag = tag
        self.__txt_string = txt_str

    def get_content(self):
        return self.__txt_string

    def get_content_trimmed_with_nl(self):
        return [x.strip() for x in self.__txt_string.splitlines()]

    def get_output_string(self):
        return self.__txt_string

    def get_tag(self):
        return self.__txt_tag


class RMTTestOutputXls:
    "Test-Class for the xlsx output class"

    def teardown_method(self):
        if self.__tmpdir:
            delete_tmp_dir(self.__tmpdir)

    def setup_method(self):
        self.__tmpdir = create_tmp_dir()
        self.oconfig = xh.default_config
        self._filename = os.path.join(self.__tmpdir, "reqs.xlsx")
        self.oconfig["output_filename"] = self._filename

    def rmttest_adding_req_header(self):
        xlsh = xh(self._filename, self.oconfig)
        xlsh.write()

        twb = openpyxl.load_workbook(filename=self._filename)
        rws = twb['Requirements']

        i = 0
        for tval in xh.default_config['headers']:
            i += 1
            assert rws.cell(row=1, column=i).value == tval
        assert i == 9
        assert rws['B1'].value == "Name"

    def rmttest_adding_req(self):
        xlsh = xh(self._filename, self.oconfig)
        xlsh.add_req(create_req(u'SW-101'))
        xlsh.add_req(create_req(u'SW-102'))
        xlsh.write()

        twb = openpyxl.load_workbook(filename=self._filename)
        rws = twb['Requirements']
        assert rws['A2'].value == "SW-101"
        assert rws['G2'].value == "007"
        assert rws['I2'].value.date().isoformat() == "1970-01-01"

        assert rws['A3'].value == "SW-102"

    def rmttest_adding_req_with_missing_field(self):
        xlsh = xh(self._filename, self.oconfig)
        req = create_req(u'SW-101')
        del(req.values['Invented by'])
        for i in range(len(req.record)):
            if req.record[i].get_content() == 'Q':
                req.record.pop(i)
                break
        with pytest.raises(AssertionError):
            xlsh.add_req(req)

    def rmttest_adding_topic(self):
        xlsh = xh(self._filename, self.oconfig)
        topic_tags = [Mock(**{'get_tag.return_value': "asdf",
                              'get_content.return_value': "qwer"})]
        topic_cfg = {'get_tags.return_value': topic_tags}
        topic = Mock(**topic_cfg)
        topic.name = "SuperTopic"
        xlsh.add_topic(topic)
        xlsh.write()

        twb = openpyxl.load_workbook(filename=self._filename)
        rws = twb['Topics']
        assert rws['A1'].value == "SuperTopic"
        assert rws['B1'].value == "asdf"
        assert rws['C1'].value == "qwer"


class RMTTestOutputXlsTemplate:
    '''Test-Class for the xlsx output with a template workbook'''

    def teardown_method(self):
        if self.__tmpdir:
            delete_tmp_dir(self.__tmpdir)

    def setup_method(self):
        self.__tmpdir = create_tmp_dir()
        self.oconfig = xh.default_config
        self._filename = os.path.join(self.__tmpdir, "reqs.xlsx")
        self.oconfig["output_filename"] = self._filename
        self.oconfig["template_filename"] = os.path.join(
            LDIR, 'DefaultTemplate.xlsx')

    def rmttest_adding_req(self):
        xlsh = xh(self._filename, self.oconfig)
        xlsh.add_req(create_req(u'SW-101'))
        xlsh.add_req(create_req(u'SW-102'))
        xlsh.write()

        twb = openpyxl.load_workbook(filename=self._filename)
        rws = twb['Requirements']
        assert rws['A3'].value == "SW-101"
        assert rws['G3'].value == "007"
        assert rws['I3'].value.date().isoformat() == "1970-01-01"

        assert rws['A4'].value == "SW-102"
